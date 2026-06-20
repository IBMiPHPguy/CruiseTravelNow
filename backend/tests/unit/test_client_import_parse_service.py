from io import BytesIO

from openpyxl import Workbook

from app.services.client_import_parse_service import (
    PREVIEW_ROW_LIMIT,
    parse_client_import_upload,
    suggest_client_import_mapping,
)
from app.services.client_import_template_service import (
    CLIENT_IMPORT_COLUMNS,
    build_client_import_template_xlsx,
)


def test_parse_template_xlsx_uses_clients_sheet():
    content = build_client_import_template_xlsx()
    result = parse_client_import_upload(content, "clients.xlsx")

    assert result["sheet_name"] == "Clients"
    assert result["source_columns"] == [column[0] for column in CLIENT_IMPORT_COLUMNS]
    assert len(result["preview_rows"]) >= 2
    assert result["preview_rows"][0][0] == "Alex"
    assert result["suggested_mapping"]["first_name"] == "first_name"
    assert result["suggested_mapping"]["last_name"] == "last_name"
    assert len(result["target_fields"]) == len(CLIENT_IMPORT_COLUMNS)


def test_parse_csv_with_aliases():
    csv_content = (
        "First Name,Last Name,Email\n"
        "Taylor,Smith,taylor@example.com\n"
        "Casey,Jones,casey@example.com\n"
    ).encode("utf-8")
    result = parse_client_import_upload(csv_content, "clients.csv")

    assert result["sheet_name"] is None
    assert result["source_columns"] == ["First Name", "Last Name", "Email"]
    assert result["suggested_mapping"]["first_name"] == "First Name"
    assert result["suggested_mapping"]["last_name"] == "Last Name"
    assert result["suggested_mapping"]["email"] == "Email"
    assert len(result["preview_rows"]) == 2


def test_suggest_mapping_avoids_duplicate_source_columns():
    mapping = suggest_client_import_mapping(["first_name", "firstname"])
    assert mapping["first_name"] == "first_name"
    assert mapping["last_name"] is None


def test_parse_xlsx_limits_preview_rows():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clients"
    headers = [column[0] for column in CLIENT_IMPORT_COLUMNS]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header)
    for row_index in range(2, PREVIEW_ROW_LIMIT + 4):
        sheet.cell(row=row_index, column=1, value=f"Client{row_index}")

    buffer = BytesIO()
    workbook.save(buffer)
    result = parse_client_import_upload(buffer.getvalue(), "many-rows.xlsx")

    assert len(result["preview_rows"]) == PREVIEW_ROW_LIMIT


def test_parse_rejects_unsupported_extension():
    try:
        parse_client_import_upload(b"not a spreadsheet", "clients.txt")
    except ValueError as exc:
        assert ".csv" in str(exc)
    else:
        raise AssertionError("Expected ValueError for unsupported extension")
