from io import BytesIO

from openpyxl import load_workbook

from app.services.client_import_template_service import (
    CLIENT_IMPORT_COLUMNS,
    CLIENT_IMPORT_TEMPLATE_FILENAME,
    build_client_import_template_xlsx,
)


def test_client_import_template_workbook_structure():
    content = build_client_import_template_xlsx()
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["Instructions", "Clients", "Field Reference"]

    clients = workbook["Clients"]
    headers = [clients.cell(row=1, column=index).value for index in range(1, len(CLIENT_IMPORT_COLUMNS) + 1)]
    assert headers == [column[0] for column in CLIENT_IMPORT_COLUMNS]
    assert clients.cell(row=2, column=1).value == "Alex"
    assert clients.cell(row=3, column=12).value == "55+ (Senior)|Educator"

    instructions = workbook["Instructions"]
    assert "client migration template" in str(instructions["A1"].value).lower()

    assert CLIENT_IMPORT_TEMPLATE_FILENAME.endswith(".xlsx")
