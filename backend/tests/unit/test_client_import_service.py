from io import BytesIO

from openpyxl import Workbook

from app.models import Passenger
from app.services.client_import_parse_service import PREVIEW_ROW_LIMIT
from app.services.client_import_template_service import build_client_import_template_xlsx
from app.services.client_import_service import execute_client_import, validate_client_import_mapping


def _full_mapping() -> dict[str, str]:
    return {
        "first_name": "first_name",
        "last_name": "last_name",
        "email": "email",
        "phone": "phone",
        "date_of_birth": "date_of_birth",
        "address_line_1": "address_line_1",
        "address_line_2": "address_line_2",
        "city": "city",
        "state_or_province": "state_or_province",
        "postal_code": "postal_code",
        "country": "country",
        "qualifiers": "qualifiers",
        "is_active": "is_active",
    }


def test_execute_client_import_from_template(db, test_user):
    content = build_client_import_template_xlsx()
    mapping = _full_mapping()

    result = execute_client_import(
        db,
        content=content,
        filename="clients.xlsx",
        mapping=mapping,
        created_by_id=test_user.id,
    )

    assert result["imported_count"] == 2
    assert result["skipped_count"] == 0
    assert result["errors"] == []

    passengers = db.query(Passenger).order_by(Passenger.id.asc()).all()
    assert len(passengers) == 2
    assert passengers[0].first_name == "Alex"
    assert passengers[0].qualifiers == ["Military"]
    assert passengers[1].last_name == "Nguyen"
    assert passengers[1].qualifiers == ["55+ (Senior)", "Educator"]
    assert passengers[1].is_active is True


def test_validate_client_import_mapping_requires_first_and_last_name():
    try:
        validate_client_import_mapping({"first_name": "first_name"})
    except ValueError as exc:
        assert "last name" in str(exc).lower()
    else:
        raise AssertionError("Expected ValueError for missing last_name mapping")


def test_execute_client_import_reports_row_errors_without_commit(db, test_user):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clients"
    headers = ["first_name", "last_name", "email"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header)
    sheet.cell(row=2, column=1, value="Valid")
    sheet.cell(row=2, column=2, value="Client")
    sheet.cell(row=2, column=3, value="valid@example.com")
    sheet.cell(row=3, column=1, value="Missing")
    sheet.cell(row=3, column=2, value="")

    buffer = BytesIO()
    workbook.save(buffer)
    mapping = {
        "first_name": "first_name",
        "last_name": "last_name",
        "email": "email",
    }

    result = execute_client_import(
        db,
        content=buffer.getvalue(),
        filename="clients.xlsx",
        mapping=mapping,
        created_by_id=test_user.id,
    )

    assert result["imported_count"] == 1
    assert len(result["errors"]) == 1
    assert result["errors"][0]["row_number"] == 3
    assert result["errors"][0]["record_label"] == "Missing"
    assert db.query(Passenger).count() == 1


def test_parse_preview_still_limited():
    from app.services.client_import_parse_service import parse_client_import_upload

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clients"
    sheet.cell(row=1, column=1, value="first_name")
    sheet.cell(row=1, column=2, value="last_name")
    for row_index in range(2, PREVIEW_ROW_LIMIT + 4):
        sheet.cell(row=row_index, column=1, value=f"Client{row_index}")
        sheet.cell(row=row_index, column=2, value="Example")

    buffer = BytesIO()
    workbook.save(buffer)
    result = parse_client_import_upload(buffer.getvalue(), "many-rows.xlsx")

    assert len(result["preview_rows"]) == PREVIEW_ROW_LIMIT
