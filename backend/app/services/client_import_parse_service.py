from __future__ import annotations

import csv
import re
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from app.services.client_import_template_service import CLIENT_IMPORT_COLUMNS

PREVIEW_ROW_LIMIT = 5

_COLUMN_ALIASES: dict[str, str] = {
    "firstname": "first_name",
    "fname": "first_name",
    "givenname": "first_name",
    "lastname": "last_name",
    "lname": "last_name",
    "surname": "last_name",
    "familyname": "last_name",
    "emailaddress": "email",
    "mail": "email",
    "phonenumber": "phone",
    "mobile": "phone",
    "cell": "phone",
    "dob": "date_of_birth",
    "birthdate": "date_of_birth",
    "dateofbirth": "date_of_birth",
    "address1": "address_line_1",
    "addressline1": "address_line_1",
    "street": "address_line_1",
    "address2": "address_line_2",
    "addressline2": "address_line_2",
    "state": "state_or_province",
    "province": "state_or_province",
    "zip": "postal_code",
    "zipcode": "postal_code",
    "postalcode": "postal_code",
    "postcode": "postal_code",
    "active": "is_active",
    "status": "is_active",
}


def _normalize_column_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", name.lower())


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _build_target_fields() -> list[dict[str, Any]]:
    return [
        {
            "field_name": field_name,
            "required": required_label == "Required",
            "description": description,
        }
        for field_name, required_label, description in CLIENT_IMPORT_COLUMNS
    ]


def suggest_client_import_mapping(source_columns: list[str]) -> dict[str, str | None]:
    normalized_sources: dict[str, str] = {}
    for column in source_columns:
        normalized = _normalize_column_name(column)
        if normalized and normalized not in normalized_sources:
            normalized_sources[normalized] = column

    mapping: dict[str, str | None] = {}
    used_sources: set[str] = set()

    for field_name, _, _ in CLIENT_IMPORT_COLUMNS:
        candidates: list[str] = []
        normalized_field = _normalize_column_name(field_name)
        if normalized_field in normalized_sources:
            candidates.append(normalized_sources[normalized_field])

        for normalized_source, original_source in normalized_sources.items():
            alias_target = _COLUMN_ALIASES.get(normalized_source)
            if alias_target == field_name and original_source not in candidates:
                candidates.append(original_source)

        chosen: str | None = None
        for candidate in candidates:
            if candidate not in used_sources:
                chosen = candidate
                break

        mapping[field_name] = chosen
        if chosen:
            used_sources.add(chosen)

    return mapping


def _parse_csv(content: bytes) -> tuple[str | None, list[str], list[list[str]]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        raise ValueError("The uploaded file has no readable rows.")

    headers = [cell.strip() for cell in rows[0]]
    if not any(headers):
        raise ValueError("The uploaded file is missing a header row.")

    data_rows: list[list[str]] = []
    for row in rows[1:]:
        values = [row[index].strip() if index < len(row) else "" for index in range(len(headers))]
        if any(values):
            data_rows.append(values)

    return None, headers, data_rows


def _choose_xlsx_sheet(workbook: Any) -> Any:
    if "Clients" in workbook.sheetnames:
        return workbook["Clients"]

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet.max_row and sheet.max_column:
            return sheet

    return workbook.active


def _parse_xlsx(content: bytes) -> tuple[str | None, list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to parse client import spreadsheets.") from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = _choose_xlsx_sheet(workbook)
        sheet_name = sheet.title

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("The uploaded spreadsheet is missing a header row.")

        headers = [_cell_to_str(value) for value in header_row]
        if not any(headers):
            raise ValueError("The uploaded spreadsheet is missing a header row.")

        data_rows: list[list[str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = [_cell_to_str(value) for value in row[: len(headers)]]
            while len(values) < len(headers):
                values.append("")
            if any(values):
                data_rows.append(values)

        return sheet_name, headers, data_rows
    finally:
        workbook.close()


def read_client_import_sheet(content: bytes, filename: str) -> tuple[str | None, list[str], list[list[str]]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return _parse_csv(content)
    if lower_name.endswith(".xlsx"):
        return _parse_xlsx(content)
    raise ValueError("Upload a .csv or .xlsx file.")


def parse_client_import_upload(content: bytes, filename: str) -> dict[str, Any]:
    sheet_name, source_columns, data_rows = read_client_import_sheet(content, filename)

    return {
        "filename": filename,
        "sheet_name": sheet_name,
        "source_columns": source_columns,
        "preview_rows": data_rows[:PREVIEW_ROW_LIMIT],
        "target_fields": _build_target_fields(),
        "suggested_mapping": suggest_client_import_mapping(source_columns),
    }
