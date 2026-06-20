from __future__ import annotations

from io import BytesIO
from typing import Any

from app.branding import BRAND_NAME
from app.constants import QUALIFIERS

CLIENT_IMPORT_TEMPLATE_FILENAME = f"{BRAND_NAME}-Client-Migration-Template.xlsx"

CLIENT_IMPORT_COLUMNS: list[tuple[str, str, str]] = [
    ("first_name", "Required", "Client first name"),
    ("last_name", "Required", "Client last name"),
    ("email", "Optional", "Primary email address"),
    ("phone", "Optional", "Phone number"),
    ("date_of_birth", "Optional", "Birth date as YYYY-MM-DD"),
    ("address_line_1", "Optional", "Street address line 1"),
    ("address_line_2", "Optional", "Street address line 2"),
    ("city", "Optional", "City"),
    ("state_or_province", "Optional", "State or province"),
    ("postal_code", "Optional", "Postal or ZIP code"),
    ("country", "Optional", "Country"),
    ("qualifiers", "Optional", "Pipe-separated values, e.g. Military|Educator"),
    ("is_active", "Optional", "Yes or No (defaults to Yes when blank)"),
]

CLIENT_IMPORT_SAMPLE_ROWS: list[dict[str, str]] = [
    {
        "first_name": "Alex",
        "last_name": "Rivera",
        "email": "alex.rivera@example.com",
        "phone": "5551234567",
        "date_of_birth": "1978-04-12",
        "address_line_1": "128 Harbor View Dr",
        "address_line_2": "Apt 4B",
        "city": "Tampa",
        "state_or_province": "FL",
        "postal_code": "33602",
        "country": "United States",
        "qualifiers": "Military",
        "is_active": "Yes",
    },
    {
        "first_name": "Jordan",
        "last_name": "Nguyen",
        "email": "jordan.nguyen@example.com",
        "phone": "5559876543",
        "date_of_birth": "1965-11-03",
        "address_line_1": "45 Lake Shore Rd",
        "address_line_2": "",
        "city": "Seattle",
        "state_or_province": "WA",
        "postal_code": "98101",
        "country": "United States",
        "qualifiers": "55+ (Senior)|Educator",
        "is_active": "Yes",
    },
]


def _autosize_columns(sheet: Any, *, max_width: int = 42) -> None:
    from openpyxl.utils import get_column_letter

    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(width + 2, 12), max_width)


def _build_instructions_sheet(sheet: Any, *, title_font: Any, body_font: Any) -> None:
    from openpyxl.styles import Alignment, Font

    sheet.title = "Instructions"
    sheet["A1"] = f"{BRAND_NAME} client migration template"
    sheet["A1"].font = title_font
    sheet["A3"] = "How to use this workbook"
    sheet["A3"].font = Font(color="1864AB", bold=True)
    instructions = [
        "1. Enter one client per row on the Clients sheet.",
        "2. Keep the header row unchanged so columns can be mapped during import.",
        "3. first_name and last_name are required for every row.",
        "4. Use YYYY-MM-DD for date_of_birth.",
        "5. Separate multiple qualifiers with a pipe character (|).",
        "6. Use Yes or No in is_active. Blank means Yes.",
        "7. Delete the sample rows before uploading your real client data.",
    ]
    row = 4
    for line in instructions:
        sheet[f"A{row}"] = line
        sheet[f"A{row}"].font = body_font
        row += 1

    row += 1
    sheet[f"A{row}"] = "Allowed qualifier values"
    sheet[f"A{row}"].font = Font(color="1864AB", bold=True)
    row += 1
    for qualifier in QUALIFIERS:
        sheet[f"A{row}"] = f"• {qualifier}"
        sheet[f"A{row}"].font = body_font
        row += 1

    row += 1
    sheet[f"A{row}"] = "Database mapping"
    sheet[f"A{row}"].font = Font(color="1864AB", bold=True)
    row += 1
    sheet[f"A{row}"] = (
        "Each Clients sheet column maps directly to a SailsPipeline client record in the passengers table. "
        "The qualifiers column uses the same allowed values as request passenger qualifiers in the CRM."
    )
    sheet[f"A{row}"].font = body_font
    sheet[f"A{row}"].alignment = Alignment(wrap_text=True)
    sheet.column_dimensions["A"].width = 88


def _build_clients_sheet(sheet: Any, *, header_fill: Any, header_font: Any) -> None:
    from openpyxl.styles import Alignment

    sheet.title = "Clients"
    headers = [column[0] for column in CLIENT_IMPORT_COLUMNS]
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_index, sample in enumerate(CLIENT_IMPORT_SAMPLE_ROWS, start=2):
        for column_index, (field_name, _, _) in enumerate(CLIENT_IMPORT_COLUMNS, start=1):
            sheet.cell(row=row_index, column=column_index, value=sample.get(field_name, ""))

    sheet.freeze_panes = "A2"
    _autosize_columns(sheet)


def _build_field_reference_sheet(sheet: Any, *, header_fill: Any, header_font: Any) -> None:
    sheet.title = "Field Reference"
    reference_headers = ["Column", "Required", "Description"]
    for index, header in enumerate(reference_headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_index, (column_name, required, description) in enumerate(CLIENT_IMPORT_COLUMNS, start=2):
        sheet.cell(row=row_index, column=1, value=column_name)
        sheet.cell(row=row_index, column=2, value=required)
        sheet.cell(row=row_index, column=3, value=description)

    sheet.freeze_panes = "A2"
    _autosize_columns(sheet, max_width=52)


def build_client_import_template_xlsx() -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to generate the client import template.") from exc

    header_fill = PatternFill("solid", fgColor="1864AB")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="1864AB", bold=True, size=14)
    body_font = Font(color="102A43")

    workbook = Workbook()
    _build_instructions_sheet(workbook.active, title_font=title_font, body_font=body_font)
    _build_clients_sheet(workbook.create_sheet("Clients"), header_fill=header_fill, header_font=header_font)
    _build_field_reference_sheet(
        workbook.create_sheet("Field Reference"),
        header_fill=header_fill,
        header_font=header_font,
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
